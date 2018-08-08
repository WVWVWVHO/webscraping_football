from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
import requests
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import timeit
from pyqt5_test import Uiform
from PyQt5.QtCore import pyqtSlot
from PyQt5.QtWidgets import QApplication, QWidget,QPushButton,QLineEdit,QLabel
import sys
import os.path

class WebscrapApp(QWidget):

    # ---------------------------pyqt ui form----------------------------
    def __init__(self, parent=None):
        super(WebscrapApp, self).__init__(parent)

        self.ui = Uiform(self)
        # print('text =' + Uiform.initUI.self.line.text())


    @pyqtSlot()
    def on_click(self):

        print('Data Downloading')
        start = timeit.default_timer()

        #------------------------------------bs part--------------------
        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/67.0.3396.62 Safari/537.36'}

        url = self.ui.line.text()
        print('url = ' + url)

        html = requests.get(url, headers=headers)
        html.encoding = 'GBK'
        content = html.text

        bs = BeautifulSoup(content, 'html.parser')
        #print(bs)

        match = bs.findAll('div',{'class': ["qbx_1", "qb2_t"]})

        matchinfo = []
        for div in match:
            #print (div)
            for t in div.text.split():
                matchinfo.append(t)
        # print(matchinfo)
        print('Finish BS')

        # ----------------------selenium part for javascript----------------------
        # instantiate a chrome options object so you can set the size and headless preference
        try:
            chrome_options = Options()
            chrome_options.add_argument("--headless")
            chrome_options.add_argument("--window-size=1920x1080")

            driver = webdriver.Chrome(chrome_options=chrome_options)
            driver.implicitly_wait(30)
            driver.get(url)

            bet365 = driver.find_element_by_id(id_='tr27')
            # print(bet365.text)

            data = bet365.text.split()
            ratio = (data[1],data[2],data[3],data[4],data[5],data[7],data[8])
            for d in ratio:
                matchinfo.append(d)
            print(matchinfo)
            # print(data)
            driver.quit()
            print('Finish Selenium')

        except:
            print('JS Failed')
            pass

        #----------------------------Append to Excel----------------------

        excel = 'football01.xlsx'
        header = ('賽事', '年', 'Round', 'Seed', '主隊', '賽果', '客隊', 'seed', 'Company', '初主', '初盤', '初客', '新主', '新盤', '新客')

        if os.path.exists(excel):
            wb = load_workbook(excel)
            sheet = wb.active
            sheet.append(matchinfo)
            wb.save(excel)
        else:
            nwb = Workbook()
            nsheet = nwb.active
            nsheet.append(header)
            nsheet.append(matchinfo)
            nwb.save(excel)



        print('Data Downloaded, Please Check Excel')

        stop = timeit.default_timer()
        time = round((stop -start),2)
        print (time)
        print('Completed')

if __name__ == '__main__':
    app = QApplication(sys.argv)
    spider = WebscrapApp()
    sys.exit(app.exec_())

